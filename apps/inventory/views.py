from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views import  View
import openpyxl
from .models import PictureList, StockList, WorkerData, WorkerBatch
import csv
import codecs
from datetime import datetime
from django.core.paginator import Paginator
from django.contrib import messages
from apps.core.messages import Message
# Create your views here.


def strip_part(pn):
    if pn in [None, '', '']:
        return None
    return ''.join(e for e in pn if e.isalnum())


class Home(View):
    def get(self, request, pk=None):
        # try:
            if not pk:
                worker_data = WorkerData.objects.filter(completed=False)\
                    .order_by("batch_id", "-match_score").first()
            else:
                worker_data = WorkerData.objects.filter(id=pk).first()
            if not worker_data:
                print("now you are here")
                return render(request, "processpage.html")    
            auto_match = worker_data.part_number
            auto_match_data_list = StockList.objects.filter(part_number=auto_match)\
                .order_by("part_number", "consignment_code", "serial_number")
            paginator = Paginator(auto_match_data_list, 12)
            page = request.GET.get('page', 1)
            auto_match_data_list = paginator.page(page)
            picture_list = PictureList.objects.filter(stock_line__isnull=False)\
                    .values("stock_line")
            return render(request, 'processpage.html', context={"worker_data": worker_data,
                "auto_match_data_list": auto_match_data_list, "picture_list": picture_list})
        # except Exception as e:
        #     return HttpResponse(str(e))

    def post(self, request):
        try:
            search = request.POST.get("search")
            search_type = request.POST.get("type")
            worker_id = request.POST.get("worker_id")
            if worker_id:
                worker_data = WorkerData.objects.filter(id=worker_id).first()
            search = strip_part(search)
            stock_list = StockList.objects.all()\
                .order_by("part_number", "consignment_code", "serial_number")
            if search_type == "part":
                stock_list = stock_list.filter(part_stripped=search)
            elif search_type == "part_contains":
                stock_list = stock_list.filter(part_stripped__contains=search)
            elif search_type == "stock_serial":
                stock_list = stock_list.filter(serial_number__startswith=search)
            auto_match = worker_data.part_number
            auto_match_data_list = StockList.objects.filter(part_number=auto_match)
            paginator = Paginator(auto_match_data_list, 12)
            page = request.GET.get('page', 1)
            auto_match_data_list = paginator.page(page)
            picture_list = PictureList.objects.values_list("stock_line")
            return  render(request, "processpage.html", context={"stock_list": stock_list,
                "worker_data": worker_data,
                "auto_match_data_list": auto_match_data_list, "picture_list": picture_list})
        except Exception as e:
            return HttpResponse(str(e))
            

class ImportStockList(View):
    def get(self, request):
        try:
            return render(request, "import_stock_list.html")
        except Exception as e:
            return HttpResponse(str(e))
    def post(self, request):
        try:
            excel_file = request.FILES["excel_file"]
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.active
            data = []
            for row in worksheet.iter_rows(min_row=2):
                    stock = StockList()
                    stock.stock_id = row[0].value
                    stock.part_number = row[1].value
                    stock.part_stripped = strip_part(row[1].value)
                    stock.consignment_code = row[2].value
                    stock.consignment_code_secondary = row[3].value
                    stock.ac_eng_serial = row[4].value
                    stock.serial_number = row[5].value
                    stock.serial_number_fake = row[6].value
                    # stock.serial_number_display = row[8].value
                    stock.applicability = row[7].value
                    stock.quantity = row[8].value
                    stock.condition_code = row[9].value
                    stock.description = row[10].value
                    # stock.lot_number = row[12].value
                    # stock.lot_padded = row[13].value
                    data.append(stock)
            StockList.objects.bulk_create(data)
            messages.add_message(request, messages.SUCCESS, Message.FILE_IMPORTED)
            return redirect("home")
        except Exception as e:
            return HttpResponse(str(e))
         


class ImportPictureFile(View):
    def get(self, request):
        try:
            return render(request, "import_picture_list.html")
        except Exception as e:
            return HttpResponse(str(e))
    def post(self, request):
        try:
            excel_file = request.FILES["excel_file"]
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.active
            data = []
            for row in worksheet.iter_rows(min_row=2):
                picture = PictureList()
                picture.picture_number = row[0].value
                picture.file_name = row[1].value
                picture.category = row[2].value
                picture.sub_category = row[3].value
                picture.image_url = row[4].value
                picture.original_file_name = row[5].value
                picture.batch_name = row[6].value
                # picture.stock_line = row[6].value
                # picture.stock_line_matched = row[7].value
                data.append(picture)
            PictureList.objects.bulk_create(data)
            messages.add_message(request, messages.SUCCESS, Message.FILE_IMPORTED)
            return redirect("home")
        except Exception as e:
            return HttpResponse(str(e))


class ExportPictureResult(View):
    def get(self, request):
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "sheet1"
            data_query = PictureList.objects.all()
            header = [
                "picture_number",
                "file_name",
                "category",
                "sub_category",
                "image_url",
                "original_file_name",
                "batch_name",
                "stock_line",
                "stock_line_matched"
            ]
            sheet.append(header)
            for row in data_query:
                data = []
                data.append(row.picture_number)
                data.append(row.file_name)
                data.append(row.category)
                data.append(row.sub_category)
                data.append(row.image_url)
                data.append(row.original_file_name)
                data.append(row.batch_name)
                if row.stock_line:
                    data.append(row.stock_line.id)
                data.append(row.stock_line_matched)
                sheet.append(data)
            wb.save("Exportpicture.xlsx")
            messages.add_message(request, messages.SUCCESS, Message.FILE_SAVED)
            return redirect("home")
        except Exception as e:
            return HttpResponse(str(e))


class ExportBatchResult(View):
    def get(self, request):
        try:
            worker_batch_list = WorkerBatch.objects.all()
            return render(request, "export_worker_batch_result.html",
                context={"worker_batch_list": worker_batch_list})
        except Exception as e:
            return HttpResponse(str(e))
    def post(self, request):
        try:
            import pdb; pdb.set_trace()
            watch_obj = request.POST.get("watch_obj")
            worker_data_result = WorkerData.objects.filter(batch_id=watch_obj)
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "sheet1"
            header = [
                "picture_id",
                "batch_id",
                "hit_id",
                "hit_type_id",
                "assignment_id",
                "worker_id",
                "assignment_status",
                "image_url",
                "picture_number",
                "part_number",
                "serial_number",
                "ac_eng_serial",
                "consignment_code",
                "comments",
                "part_number_correct",
                "part_number_entered",
                "serial_number_correct",
                "serial_number_entered",
                "ac_eng_serial_correct",
                "ac_eng_serial_entered",
                "consignment_code_correct",
                "consignment_code_entered",
                "bonus_difficult_to_read",
                "bonus_inferred_info",
                "no_match_available",
                "completed",
                "accept",
                "completed_time",
                "match_score"
            ]
            sheet.append(header)
            for row in worker_data_result:
                data = []
                if row.picture_id:
                    data.append(row.picture_id.id)
                else:
                    data.append(row.picture_id)
                if row.batch_id:
                    data.append(row.batch_id.id)
                else:
                    data.append(row.batch_id)
                data.append(row.hit_id)
                data.append(row.hit_type_id)
                data.append(row.assignment_id)
                data.append(row.worker_id)
                data.append(row.assignment_status)
                data.append(row.image_url)
                data.append(row.picture_number)
                data.append(row.part_number)
                data.append(row.serial_number)
                data.append(row.ac_eng_serial)
                data.append(row.consignment_code)
                data.append(row.comments)
                data.append(row.part_number_correct)
                data.append(row.part_number_entered)
                data.append(row.serial_number_correct)
                data.append(row.serial_number_entered)
                data.append(row.ac_eng_serial_correct)
                data.append(row.ac_eng_serial_entered)
                data.append(row.consignment_code_correct)
                data.append(row.consignment_code_entered)
                data.append(row.bonus_difficult_to_read)
                data.append(row.bonus_inferred_info)
                data.append(row.no_match_available)
                data.append(row.completed)
                data.append(row.accept)
                data.append(row.completed_time)
                data.append(row.match_score)
                sheet.append(data)
            wb.save("ExportBatchResult.xlsx")
            messages.add_message(request, messages.SUCCESS, Message.FILE_SAVED)
            return redirect("home")
        except Exception as e:
            return HttpResponse(str(e))


class ImportWorkerBatch(View):
    def get(self, request):
        try:
            return render(request, "import_worker_batch.html")
        except Exception as e:
            return HttpResponse(str(e))

    def post(self, request):
        try:
            batch_name = request.POST.get("batch_name")
            file_name = request.FILES.get("csv_file")
            worker_batch_id = WorkerBatch.objects.create(
                batch_name=batch_name,
                batch_file_name=file_name.name)
            data = []
            worksheet = csv.reader(codecs.iterdecode(file_name, 'utf-8'))
            next(worksheet)
            for row in worksheet:
                picture_id = PictureList.objects.filter(picture_number=row[28]).first()
                worker_data = WorkerData()
                worker_data.picture_id = picture_id
                worker_data.batch_id = worker_batch_id
                worker_data.hit_id = row[0]
                worker_data.hit_type_id = row[1]
                worker_data.assignment_id = row[14]
                worker_data.worker_id = row[15]
                worker_data.assignment_status = row[16]
                worker_data.image_url = row[27]
                worker_data.picture_number = row[28]
                if "part_number" in eval(row[29])[0].keys():
                    worker_data.part_number = eval(row[29])[0]["part_number"]
                if "serial_number" in eval(row[29])[0].keys():
                    worker_data.serial_number = eval(row[29])[0]["serial_number"]
                if "ac_esn" in eval(row[29])[0].keys():
                    worker_data.ac_eng_serial = eval(row[29])[0]["ac_esn"]
                if "cons_code" in eval(row[29])[0].keys():
                    worker_data.consignment_code = eval(row[29])[0]["cons_code"]
                if "comments" in eval(row[29])[0].keys():
                    worker_data.comments = eval(row[29])[0]["comments"]   
                data.append(worker_data)
            WorkerData.objects.bulk_create(data)
            messages.add_message(request, messages.SUCCESS, Message.FILE_IMPORTED)
            return redirect("home")
        except Exception as e:
            return HttpResponse(str(e))


class NoMatchAvailableView(View):
    def post(self, request):
        try:
            worker_id = request.POST.get("worker_id")
            if not worker_id:
                return redirect("home")
            worker_data = WorkerData.objects.filter(id=worker_id).first()
            if worker_data:
                worker_data.no_match_available = True
                worker_data.completed = True
                worker_data.completed_time = datetime.now()
                worker_data.save()
                messages.add_message(request, messages.SUCCESS, Message.SENT_TO_NO_MATCH_AVAILABLE)
                return redirect("home" )
            return redirect("home")
        except Exception as e:
            return HttpResponse(str(e))


class MoveToEndOfQueueView(View):
    def post(self, request):
        try:
            worker_id = request.POST.get("worker_id")
            if not worker_id:
                return redirect("home")
            worker_data = WorkerData.objects.filter(id=worker_id).first()
            if worker_data:
                worker_data.match_score = 1
                worker_data.save()
                messages.add_message(request, messages.SUCCESS, Message.MOVE_TO_END_OF_QUEUE)
                return redirect("home")
            return redirect("home")
        except Exception as e:
            return HttpResponse(str(e))


class ChangeButtonView(View):
    def post(self, request):
        try:
            worker_id = request.POST.get("worker_id")
            button_type = request.POST.get("button_type")
            if not worker_id:
                return redirect("home")
            worker_data = WorkerData.objects.filter(id=worker_id).first()
            if worker_data:
                if "part_number_correct" == button_type:
                    if worker_data.part_number_correct:
                        worker_data.part_number_correct = False
                    else:
                        worker_data.part_number_correct = True
                elif "part_number_entered" == button_type:
                    if worker_data.part_number_entered:
                        worker_data.part_number_entered = False
                    else:
                        worker_data.part_number_entered = True
                elif "serial_number_correct" == button_type :
                    if worker_data.serial_number_correct:
                        worker_data.serial_number_correct = False
                    else:
                        worker_data.serial_number_correct = True
                elif "serial_number_entered" == button_type:
                    if worker_data.serial_number_entered:
                        worker_data.serial_number_entered = False
                    else:
                        worker_data.serial_number_entered = True
                elif "ac_eng_serial_correct" == button_type:
                    if worker_data.ac_eng_serial_correct:
                        worker_data.ac_eng_serial_correct = False
                    else:
                        worker_data.ac_eng_serial_correct = True
                elif "ac_eng_serial_entered" == button_type:
                    if worker_data.ac_eng_serial_entered:
                        worker_data.ac_eng_serial_entered = False
                    else:
                        worker_data.ac_eng_serial_entered = True
                elif "consignment_code_correct" == button_type:
                    if worker_data.consignment_code_correct:
                        worker_data.consignment_code_correct = False
                    else:
                        worker_data.consignment_code_correct = True 
                elif "consignment_code_entered" == button_type:
                    if worker_data.consignment_code_entered:
                        worker_data.consignment_code_entered = False
                    else:
                        worker_data.consignment_code_entered = True        
                worker_data.save()
                messages.add_message(request, messages.SUCCESS, Message.BUTTON_CHANGES)
                return redirect("home_with_id", pk=worker_data.id)
            return redirect("home_with_id", pk=worker_data.id)
        except Exception as e:
            return HttpResponse(str(e))


class DifficultToReadView(View):
    def post(self, request):
        try:
            worker_id = request.POST.get("worker_id")
            if not worker_id:
                return redirect("home")
            worker_data = WorkerData.objects.filter(id=worker_id).first()
            if worker_data:
                if worker_data.bonus_difficult_to_read:
                    worker_data.bonus_difficult_to_read = False
                else:
                    worker_data.bonus_difficult_to_read = True
                worker_data.save()
                messages.add_message(request, messages.SUCCESS, Message.BUTTON_CHANGES)
                return redirect("home_with_id", pk=worker_data.id)
            return redirect("home_with_id", pk=worker_data.id)
        except Exception as e:
            return HttpResponse(str(e))


class InferredInfoView(View):
    def post(self, request):
        try:
            worker_id = request.POST.get("worker_id")
            if not worker_id:
                return redirect("home")
            worker_data = WorkerData.objects.filter(id=worker_id).first()
            if worker_data:
                if worker_data.bonus_inferred_info:
                    worker_data.bonus_inferred_info = False
                else:
                    worker_data.bonus_inferred_info = True
                worker_data.save()
                messages.add_message(request, messages.SUCCESS, Message.BUTTON_CHANGES)
                return redirect("home_with_id", pk=worker_data.id)
            return redirect("home_with_id", pk=worker_data.id)
        except Exception as e:
            return HttpResponse(str(e))


class BatchStatusView(View):
    def get(self, request):
        try:
            batch_status = WorkerData.objects.distinct('batch_id')
            data_list = []
            for batch in batch_status:
                batch_file_name = batch.batch_id.batch_file_name
                total_rows = WorkerData.objects.filter(batch_id__batch_file_name=batch_file_name).count()
                completed_rows = WorkerData.objects.filter(batch_id__batch_file_name=batch_file_name, completed=True).count()
                remaining_rows = WorkerData.objects.filter(batch_id__batch_file_name=batch_file_name, completed=False).count()
                obj = {"batch_file_name": batch_file_name,
                    "total_rows": total_rows,
                    "completed_rows": completed_rows,
                    "remaining_rows": remaining_rows}
                data_list.append(obj)
            return render(request, "batch_status.html", context={"data_list": data_list})
        except Exception as e:
            return HttpResponse(str(e))


class RecentMatchesView(View):
    def get(self, request):
        try:
            recent_matches = WorkerData.objects.filter(completed=True, no_match_available=False).order_by("-completed_time")
            return render(request, "recent_match.html", context={"recent_matches": recent_matches})
        except Exception as e:
            return HttpResponse(str(e))

    def post(self, request):
        try:
            worker_id = request.POST.get("worker_id")
            stock_list_id = request.POST.get("stock_list_id")
            worker_data = WorkerData.objects.filter(id=worker_id).first()
            worker_data.completed = False
            worker_data.completed_time = None
            stock_data = StockList.objects.filter(id=stock_list_id).first()
            worker_data.picture_id.stock_line = None
            worker_data.picture_id.stock_line_matched = False
            worker_data.save()
            worker_data.picture_id.save()
            return redirect("recent-matches")
        except Exception as e:
            return HttpResponse(str(e))


class StockIdBtnClickView(View):
    def post(self, request):
        try:
            worker_id = request.POST.get("worker_id")
            stock_list_id = request.POST.get("stock_list_id")
            worker_data = WorkerData.objects.filter(id=worker_id).first()
            worker_data.completed = True
            worker_data.completed_time = datetime.now()
            stock_data = StockList.objects.filter(id=stock_list_id).first()
            worker_data.picture_id.stock_line = stock_data
            worker_data.picture_id.stock_line_matched = True
            worker_data.save()
            worker_data.picture_id.save()
            return redirect("home")
        except Exception as e:
            return HttpResponse(str(e))

